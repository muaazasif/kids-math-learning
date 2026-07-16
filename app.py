from flask import Flask,render_template,request,send_file,jsonify
from openpyxl import Workbook,load_workbook
import os
app=Flask(__name__)
BASE=os.path.dirname(__file__)
DATA=os.path.join(BASE,"data")
os.makedirs(DATA,exist_ok=True)
F=os.path.join(DATA,"results.xlsx")
@app.route("/")
def home():
    return render_template("index.html")
@app.route("/save",methods=["POST"])
def save():
    d=request.get_json()
    ans=int(d.get("answer",0))
    correct=5
    marks=100 if ans==correct else 0
    wb=load_workbook(F) if os.path.exists(F) else Workbook()
    ws=wb.active
    if ws.max_row==1 and ws["A1"].value is None:
        ws.append(["Question","Answer","Correct","AI Marks"])
    ws.append(["2+3",ans,ans==correct,marks]); wb.save(F)
    return jsonify({"marks":marks})
@app.route("/download")
def download():
    return send_file(F,as_attachment=True)
